# -*- coding: utf-8 -*-
"""
Created on Fri Feb  7 11:28:15 2025

@author: evang
"""
import re

import pandas as pd
import os
import numpy as np
from openpyxl import load_workbook
from sympy.codegen.ast import continue_


class merge_tb(object):
    def __init__(self,path_tb,path):
        self.path_tb=path_tb
        self.path=path
    def change_name(self):
        ls_old=os.listdir(self.path_tb)
        # print(len(ls_old[0]))
        # 需要弄个开关用“_”识别
        if len(ls_old[0])>15:

            for i in range(len(ls_old)):
                old_name=self.path_tb+ls_old[i]
                new_name=self.path_tb+ls_old[i].split('_')[0]+'.xlsx'
                os.rename(old_name,new_name)
        else:
            print(ls_old[0])
    def chuli(self):
        self.change_name()
        df_baoxiao=pd.read_excel(self.path_tb+'费用报销单.xlsx')
        df_baoxiao=df_baoxiao.loc[:,['单据编号','单据状态']]
        df_yingfu = pd.read_excel(self.path_tb + '应付单.xlsx')
        df_yingfu = df_yingfu.loc[:, ['单据编号', '单据状态']]
        df_all=pd.concat([df_baoxiao,df_yingfu],axis=0)
        df_all.dropna(inplace=True)
        # df_all.to_excel('.//123.xlsx')
    #     ===========================================
    #     处理信息中心表格
        df_xinxi=pd.read_excel(self.path_tb+'信息中心.xlsx')
        df_xinxi['流程时长'] = np.nan
        for i in range(len(df_xinxi['发起时间'])):
            df_xinxi['发起时间'][i]=pd.Timestamp(df_xinxi['发起时间'][i])
            df_xinxi['任务到达时间'][i] = pd.Timestamp(df_xinxi['任务到达时间'][i])
            df_xinxi['流程时长'][i] = df_xinxi['任务到达时间'][i] - df_xinxi['发起时间'][i]
            df_xinxi['流程时长'][i] = df_xinxi['流程时长'][i].days

        # df_xinxi['流程时长']=df_xinxi['任务到达时间']-df_xinxi['发起时间']
        # df_xinxi['流程时长']=df_xinxi['流程时长'].days

        # df_xinxi = df_xinxi.loc[:, ['流程实例', '当前处理人']]
        df_xinxi=df_xinxi.loc[:,['流程实例','当前处理人','流程时长','任务到达时间']]
        df_xinxi['单号']=np.nan
        for i in range(len(df_xinxi['流程实例'])):
            df_xinxi['单号'][i]=df_xinxi['流程实例'][i].split('_')[0]


        df=pd.merge(df_all,df_xinxi,left_on='单据编号',right_on='单号',how='left')
        df = df.loc[:, ['单据编号','单据状态','当前处理人','任务到达时间']] #此处修改
        for i in range(len(df['单据编号'])):
            if df['单据状态'][i]=='审核中':
                df['单据状态'][i]=df['当前处理人'][i]
            else:
                df['单据状态']=df['单据状态']
        # df.to_excel('.//sdf.xlsx',index=False)



    # 合并表格填充同步data页面
    #!!!!!!!!!!!!!!!!!注意导入文件时候如果不加keep_vba=True，会破坏VBA文件导致文件损坏
        wb = load_workbook(self.path,keep_vba=True)
        ws=wb['data']
        ws1=wb['系统汇总页面']
        max_row=ws1.max_row
        # 填充后面合成的df 数据
        for i in range(len(df.columns)):#获取列数据
            for j in range(len(df)):#获取行数据
                ws.cell(row=j+1,column=i+1).value=df.iloc[j,i]


        # 填充超链接
        link='D:\\工作文件\\系统填报单据归档\\'#增加超级链接
        for i in range(2,max_row):
            minxi_nb=ws1.cell(row=i,column=3)
            gongsi_nb=ws1.cell(row=i,column=11)

            link_info=link+str(gongsi_nb.value)
            ls_info=os.listdir(link_info)
            # print(ls_info)



            # for j in ls_info:
            #     if minxi_nb==re.match(minxi_nb,j).group():
            #         minxi_nb.hyperlink = (link + gongsi_nb.value+j)


            minxi_nb.hyperlink=(link+gongsi_nb.value)



        wb.save('.//系统填报汇总.xlsm')

    def run(self):
        pass

if __name__=='__main__':
    path = './/系统填报汇总.xlsm'
    path_tb='.//系统导出//'
    merge_tb(path_tb,path).chuli()
    